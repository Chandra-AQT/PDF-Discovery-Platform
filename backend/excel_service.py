import os
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_excel(pdf_links: list, folder: str) -> str:
    excel_path = os.path.join(folder, "pdf_links.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Links"

    # Header style
    header_fill = PatternFill("solid", fgColor="0284C7")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ["#", "Filename", "Domain", "PDF URL"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, link in enumerate(pdf_links, start=1):
        parsed   = urlparse(link)
        filename = link.split("/")[-1].split("?")[0].strip() or "document.pdf"
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=filename)
        ws.cell(row=i+1, column=3, value=parsed.netloc)
        ws.cell(row=i+1, column=4, value=link)

    # Auto column widths
    col_widths = [6, 40, 30, 80]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(excel_path)
    return excel_path
